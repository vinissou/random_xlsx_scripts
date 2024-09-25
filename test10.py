import os
import openpyxl



class XLSXrow:
    A = '' 
    B = ''
    C = ''
    D = ''
    E = ''
    F = ''
    G = ''
    H = ''


def reset_xlsx():
    if os.name == 'nt':
        os.system("del out.xlsx")
    else:
        os.system("rm out.xlsx")



def read_specific_cells(file_path, out_path):
    input_excel  = openpyxl.load_workbook(file_path) 
    sheet        = input_excel.active

    output_excel = openpyxl.Workbook()
    NewXLSX = XLSXrow()
    sheet_out1   = output_excel.create_sheet('Prontuários', 0)

    sess = 0;
    extract_sess = 0

    print('')

    for row in sheet:
        for cell in row:
            if cell.column == 1:
                sess = cell.value
                #print (cell.column)
                #print (cell.row)
                sheet_out1[F'A{cell.row}'] = sess

            if cell.column == 2:
                coord = cell.coordinate
                val   = cell.value

                print(F"\tProcessando a célula: {coord}", end= "\r")

                if isinstance(val, int):
                    digit = int(repr(val)[-1])
                    pront = int(cell.value/10)

                    if pront > 9:
                        extract_sess = int(str(pront)[-2:])
                    else:
                        extract_sess = pront

                    if extract_sess != sess:
                        print("\n SESSÃO ERRADA =  ", extract_sess)
                        exit()
                else:
                    print("NOT INT ---- CORD:", coord)
                    exit()

                sheet_out1[F'B{cell.row}'] = pront
                sheet_out1[F'C{cell.row}'] = digit

            if cell.column == 3:
                sheet_out1[F'D{cell.row}'] = cell.value

    output_excel.save(out_path)


#reset_xlsx()
file_path = 'data.xlsx'
out_path = 'out.xlsx'
read_specific_cells(file_path, out_path)
print ("\n==== COMPLETO =====\n")
