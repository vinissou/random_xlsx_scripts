import os
import openpyxl
import numpy
import pandas 



class Caixas: 
    HP = ''
    HM = ''
    GE = ''
    FU = ''
    OB = ''



def reset_xlsx():
    if os.name == 'nt':
        os.system("del out.xlsx")
    else:
        os.system("rm out.xlsx")


def ArrayXLSX(input_file):
    file = pandas.read_excel(input_file, sheet_name=0)
    array = file.to_numpy()
    return array

def CheckCaixa(array, pront):
    for x in array:
        if (x[0] == 35):
            if (x[3] == pront):
                print('begin match')
            if (x[3] < pront and x[4] > pront ):
                print('range match!!!!')

    #return array


# TODO all this ↓ 
#Talvez separar por seções? 
# seção_atual == array[1>div>exact>range.]
#
# comparando a sesão primeiro vai ser mais rápido
# 
#  if: sessão match
#      for: os números exatos
#      elif: ranges 
#      
#
#def init_caixas(file_caixas):
#    caixas     = openpyxl.load_workbook(file_caixas)
#    sheetnames = caixas.sheetnames 
#
#
#    for i in range(1, 5, 1)
#        sheet = caixas[sheetnames[i]]
#
#        for row in sheet
#            for cell in row:
#                if not isinstance(val, int):
#                    print("NOT INT >>>> CORD:", cell.coordinate)
#                    exit()
#
#                val == cell.value
#
#                if row == 1 
#                    xxx
#                if row == 2
#                    xx
#                if 
#                if i == 4
#                    continue
#
#                if row == 4
#                
#                if row == 6
#                   #separa os números
#
#
#
#def CheckCaixa(seção, pront, divisão):
#    return_caixas = Caixas()
# #check com o array   
#    Caixas.HP = matrix[x. x. .x]
#
#
#
#def read_specific_cells(file_caixas, out_path):
#    nomes_excel  = openpyxl.load_workbook(file_nomes) 
#    sheet        = nomes_excel.active
#
#    output_excel = openpyxl.Workbook()
#    sheet_out1   = output_excel.create_sheet('Prontuários', 0)
#
#    seção = 0
#    pront = 0
#
#    print('')
#
#    for row in sheet:
#        print(F"\tProcessando:\t" cell.row, end= "\r")
#
#        if row == 1: 
#            sheet_out1['A1'] = "Sessão"
#            sheet_out1['B1'] = "Prontuário"
#            sheet_out1['C1'] = "Dígito"
#            sheet_out1['D1'] = "Paciente"
#            sheet_out1['E1'] = "HCPA"
#            sheet_out1['F1'] = "Humaitá"
#            sheet_out1['G1'] = "Genética"
#            sheet_out1['H1'] = "Funcionários"
#            sheet_out1['I1'] = "Observações"
#            continue
#
#        for cell in row:
#            val   = cell.value
#
#            if cell.column == 1: #sessão
#                seção = val
#                sheet_out1[F'A{cell.row}'] = seção
#
#            elif cell.column == 2: #pront
#                pront = val
#                sheet_out1[F'B{cell.row}'] = pront
#
#            elif cell.column == 3: #digit
#                sheet_out1[F'C{cell.row}'] = val
#
#            elif cell.column == 4: #nome
#                sheet_out1[F'D{cell.row}'] = pront
#
#            elif cell.column == 5:
#                HCPA = CheckCaixa(seção, pront, 1)
#                sheet_out1[F'D{cell.row}'] = HCPA
#
#            elif cell.column == 6:
#                Humaitá = CheckCaixa(seção, pront, 2)
#                sheet_out1[F'D{cell.row}'] = Humaitá
#
#            elif cell.column == 7:
#                Genética = CheckCaixa(seção, pront, 3)
#                sheet_out1[F'D{cell.row}'] = Genética
#
#            elif cell.column == 8:
#                Funcionários = CheckCaixa(seção, pront, 4)
#                sheet_out1[F'D{cell.row}'] = Funcionários
#
#            elif cell.column == 9:
#                Observações = CheckCaixa(seção, pront, 5)
#                sheet_out1[F'D{cell.row}'] = Observações
#
#    output_excel.save(out_path)



#reset_xlsx()

#caixas_array = init_caixas(file_caixas)
#read_specific_cells(file_caixas, out_path)


file_nomes = 'nomes.xlsx'
file_caixas = 'caixas.xlsx'
out_path = 'out.xlsx'
array = ArrayXLSX(file_caixas)
CheckCaixa(array, 830000)

print ("\n==== COMPLETO =====\n")


#if isinstance(val, int):
#else:
#    print("NOT INT ---- CORD:", coord)
#    exit()
